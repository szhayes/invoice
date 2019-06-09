#!/usr/bin/env python3
import argparse
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as PXYLImage
from datetime import datetime, timedelta
import os, PIL
from PIL import Image # as PILImage
import yagmail

basedir = '/mnt/c/Users/shayes01'
imgdir = basedir + '/Pictures/ApprovedTime/'
wbdir = basedir + '/Downloads/'

# parse command line
parser = argparse.ArgumentParser()
parser.add_argument('-H', '--Hours', type=float, default=75, help="modify hours billed")
parser.add_argument('-t', '--tes', action='store_true', help="TES mode send invoice to TES")
parser.add_argument('-m', '--mail', action='store_true', help="Send Mail")
args = parser.parse_args()

# update worksheet with current values
wb = load_workbook(filename = wbdir + 'invoice-template.xlsx')
ws = wb.active

now = datetime.today()
ws['D3'] = now.date()
ws['D4'] = 'CSZH-'+str(now.date())
ws['A19'] = " - for two week period ending " + str(now.date() + timedelta(1))
if args.Hours != 75:
    ws['B18'] = args.Hours

# resize images & add to invoice
src_images = []
src_files = []
for filename in os.listdir(imgdir):
    if filename.endswith('.JPG'):
        src_images.append(imgdir + filename)
        src_files.append(filename)
print(src_images)
new_images = []
cells = ['A52','A67']
basewidth = 690

for i in src_images:
    foo = src_images.index(i)
    img = PIL.Image.open(i)
    wpercent = (basewidth / float(img.size[0]))
    hsize = int((float(img.size[1]) * float(wpercent)))
    img = img.resize((basewidth, hsize), PIL.Image.ANTIALIAS)
    new_images.append(i[:-4]+'_resized'+i[-4:])
    # print(foo, new_images[foo])
    img = img.save(new_images[foo])
    img2 = PXYLImage(new_images[foo])
    ws.add_image(img2, cells[foo])

invoice = wbdir + "/HAYES invoice-" + str(now.date()) + ".xlsx"
wb.save(invoice)

for files in new_images:
    os.remove(files)

for files in src_files:
    os.rename(f'{imgdir}{files}', f'{imgdir}Submitted/{files}') 


# setup mail
if args.mail:
    m_recipient = ['szhayes@gmail.com']
    if args.tes:
        m_recipient.append('payroll@tes.net')
    m_subject = f'Invoice for Hayes @ Teranet for period ending {now.date()}'
    m_body = f"Please find attached the invoice for Stuart Hayes @ Teranet\n\nFor the period ending {now.date()}\n\nRegards\nStuart Hayes\nC. 416.357.4525"

    print('Sending mail ...')
    #send email via gmail
    yag = yagmail.SMTP('szhayes@gmail.com')
    yag.send(
        to=m_recipient,
        subject=m_subject,
        contents=m_body,
        attachments=invoice
        )
